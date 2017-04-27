#!/usr/bin/env python

# Set of data management interfaces

import sys
import os
import logging
import datetime
import math

import pandas as pd
import numpy as np
import openpyxl as op
from openpyxl.utils.dataframe import dataframe_to_rows

def is_leap_year(year_to_test):
    """Test if year is a leap year
    
        Args:
            year_to_test: year to test

        Returns:
            boolean: True or False
        """
    if year_to_test % 4 == 0 and year_to_test %100 != 0 or year_to_test % 400 == 0:
        return True
    else:
        return False

def get_ts_freq(time_step, ts_quantity, wyem = 12):
    """ Get pandas time frequency given time_step and ts_quantity
    
     Args:
        time_step: RiverWare style string timestep
        ts_quantity: Interger number of time_steps's in interval

    Returns:
        ts_freq: pandas time frequency
    """
    ts_freq = None
    if time_step == 'day':
        ts_freq = "D"
    elif time_step == 'year':
        ts_freq = water_year_agg_func(wyem)
    elif time_step == 'month':
        ts_freq = "M"
    elif time_step == 'hour':
        if ts_quantity == 1:
            ts_freq = "H"
        else:
            ts_freq = str(ts_quantity) + "H"
    elif time_step == 'minute':
        if ts_quantity == 1:
            ts_freq = "T"
        else:
            ts_freq = str(ts_quantity) + "T"
    elif time_step == 'week':
        ts_freq = "W"
    else:
        logging.error('\nERROR: Timestep {} and ts quantity {} are an invalid combination', format(time_step, ts_quantity))
    return ts_freq

def water_year_agg_func(wyem):
    """Sets annual aggregation function for water year end month
    Args:
        wyem: Water Year End Month
    Return; Water year aggregation function
    """
    ann_freqs = ['A-JAN', 'A-FEB', 'A-MAR', 'A-APR', 'A-MAY', 'A-JUN'] + \
                ['A-JUL', 'A-AUG', 'A-SEP', 'A-OCT', 'A-NOV', 'A-DEC']
    return ann_freqs[wyem - 1]

def make_dt_index(time_step, ts_quantity, start_dt, end_dt, wyem = 12):
    """ Make a pandas DatetimeIndex from specified dates, time_step and ts_quantity
    
     Args:
        time_step: RiverWare style string timestep
        ts_quantity: Interger number of time_steps's in interval
        start_dt: starting date time
        end_dt: ending date time
        wyem: Water Year End Month

    Returns:
        dt_index:pandas DatetimeIndex
    """
    dt_index = None
    if time_step == 'day':
        dt_index =pd.date_range(start_dt, end_dt, freq = "D", name = "date")
    elif time_step == 'year':
        dt_index = pd.date_range(start_dt, end_dt, freq = water_year_agg_func(wyem), name = "date")
    elif time_step == 'month':
        dt_index = pd.date_range(start_dt, end_dt, freq = "M", name = "date")
    elif time_step == 'hour':
        if ts_quantity == 1:
            dt_index =pd.date_range(start_dt, end_dt, freq = "H", name = "date")
        else:
            dt_index = pd.date_range(start_dt, end_dt, freq = str(ts_quantity) + "H", name = "date")
    elif time_step == 'minute':
        if ts_quantity == 1:
            dt_index = pd.date_range(start_dt, end_dt, freq = "T", name = "date")
        else:
            dt_index = pd.date_range(start_dt, end_dt, freq = str(ts_quantity) + "T", name = "date")
    elif time_step == 'week':
        dow = start_dt.dayofweek
        if dow == 0:
            dt_index = pd.date_range(start_dt, end_dt, freq = "W-MON", name = "date")
        elif dow == 1:
            dt_index = pd.date_range(start_dt, end_dt, freq = "W-TUE", name = "date")
        elif dow == 2:
            dt_index = pd.date_range(start_dt, end_dt, freq = "W-WED", name = "date")
        elif dow == 3:
            dt_index = pd.date_range(start_dt, end_dt, freq = "W-THU", name = "date")
        elif dow == 4:
            dt_index = pd.date_range(start_dt, end_dt, freq = "W-FRI", name = "date")
        elif dow == 5:
            dt_index = pd.date_range(start_dt, end_dt, freq = "W-SAT", name = "date")
        elif dow == 6:
            dt_index = pd.date_range(start_dt, end_dt, freq = "W-SUN", name = "date")
        else:
            dt_index = pd.date_range(start_dt, end_dt, freq = "W-SUN", name = "date")
    else:
        logging.error('\nERROR: Timestep {} and ts quantity {} are an invalid combination', format(time_step, ts_quantity))
    return dt_index

def make_ts_dataframe(time_step, ts_quantity, start_dt, end_dt, wyem = 12):
    """ Make a pandas dataframe from specified dates, time_step and ts_quantity
    
     Args:
        time_step: RiverWare style string timestep
        ts_quantity: Interger number of time_steps's in interval
        start_dt: starting date time
        end_dt: ending date time
        wyem: Water Year End Month

    Returns:
        Empty pandas datafame wihh indexed dates
    """
    ts_dataframe = None
    if time_step == 'day':
        ts_dataframe = pd.DataFrame(index = pd.date_range(start_dt, end_dt, freq = "D", name = "date"))
    elif time_step == 'year':
        ts_dataframe = pd.DataFrame(index = pd.date_range(start_dt, end_dt, freq = water_year_agg_func(wyem), name = "date"))
    elif time_step == 'month':
        ts_dataframe = pd.DataFrame(index = pd.date_range(start_dt, end_dt, freq = "M", name = "date"))
    elif time_step == 'hour':
        if ts_quantity == 1:
            ts_dataframe = pd.DataFrame(index = pd.date_range(start_dt, end_dt, freq = "H", name = "date"))
        else:
            ts_dataframe = pd.DataFrame(index = pd.date_range(start_dt, end_dt, freq = str(ts_quantity) + "H", name = "date"))
    elif time_step == 'minute':
        if ts_quantity == 1:
            ts_dataframe = pd.DataFrame(index = pd.date_range(start_dt, end_dt, freq = "T", name = "date"))
        else:
            ts_dataframe = pd.DataFrame(index = pd.date_range(start_dt, end_dt, freq = str(ts_quantity) + "T", name = "date"))
    elif time_step == 'week':
        dow = start_dt.dayofweek
        if dow == 0:
            ts_dataframe = pd.DataFrame(index = pd.date_range(start_dt, end_dt, freq = "W-MON", name = "date"))
        elif dow == 1:
            ts_dataframe = pd.DataFrame(index = pd.date_range(start_dt, end_dt, freq = "W-TUE", name = "date"))
        elif dow == 2:
            ts_dataframe = pd.DataFrame(index = pd.date_range(start_dt, end_dt, freq = "W-WED", name = "date"))
        elif dow == 3:
            ts_dataframe = pd.DataFrame(index = pd.date_range(start_dt, end_dt, freq = "W-THU", name = "date"))
        elif dow == 4:
            ts_dataframe = pd.DataFrame(index = pd.date_range(start_dt, end_dt, freq = "W-FRI", name = "date"))
        elif dow == 5:
            ts_dataframe = pd.DataFrame(index = pd.date_range(start_dt, end_dt, freq = "W-SAT", name = "date"))
        elif dow == 6:
            ts_dataframe = pd.DataFrame(index = pd.date_range(start_dt, end_dt, freq = "W-SUN", name = "date"))
        else:
            ts_dataframe = pd.DataFrame(index = pd.date_range(start_dt, end_dt, freq = "W-SUN", name = "date"))
    else:
        logging.error('\nERROR: Timestep {} and ts quantity {} are an invalid combination', format(time_step, ts_quantity))
    return ts_dataframe

def char_to_numeric_month(cmonth):
    """converts 3 character month to numberic month
    
    Args:
        cmonth: 3 character month
    Return:
        nmonth: numeric month
    """
    cmonths = ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']
    nmonth = cmonths.index(cmonth.upper()) + 1
    return nmonth

def ReadOneColumnSlot(file_path, header_lines, names_line, stationToRead, 
        parameterToRead, units, scaleFactor, time_step, ts_quantity, 
        valuesSeparator, start_dt = None, end_dt = None, 
        mia_value = 'NaN', wyem = 12):
    """Reads column slot data for one station and parameter
    
     Args:
        file_path: fully specified file path
        header_lines: number of header lines
        names_line: line of header names
        stationToRead: station to read
        parameterToRead: parameter to read
        units: units of parameter
        scaleFactor: scale to apply to parameter
        time_step: RiverWare style string timestep
        ts_quantity: Interger number of time_steps's in interval
        valuesSeparator: separator of values
        start_dt: starting date time
        end_dt: ending date time
        mia_value: missing value
        wyem: Water Year End Month

    Returns:
        populated dataframe
    """
    return_df = None
    lc_station = stationToRead.lower()
    lc_param = parameterToRead.lower()
    try:
        # Get list of 0 based line numbers to skip - Ignore header but assume header was set as 1's based index
        data_skip = [i for i in range(header_lines) if i + 1 <> names_line]
        input_df = pd.read_table(file_path, engine = 'python',
                header = names_line - len(data_skip) - 1, skiprows = data_skip, 
                sep = valuesSeparator, na_values = mia_value)
        if input_df.empty:
            logging.error("No data read in file" + file_path)
            return return_df
        input_columns = list(input_df.columns)
        lc_columns = [x.lower() for x in input_columns]
        
        # determine date column

        try:
            date_column = lc_columns.index('date')
        except:
            date_column = 0
        date_column_name = input_columns[date_column]
        input_columns.remove(date_column_name)
        lc_columns.remove(date_column_name.lower())
            
        # set date column as index

        input_df = input_df.rename(columns = {date_column_name:'date'})
        
        # make sure that daily, monthly and annual data use end of period dates and do not include a time stamp
        
        input_df['date'] = pd.to_datetime(input_df['date'])
        input_df.set_index('date', inplace = True)
        if time_step == 'day' or time_step == 'month' or time_step == 'year':
            input_df['year'] = input_df.index.year
            input_df['month'] = input_df.index.month
            if time_step == 'day':
                input_df['day'] = input_df.index.day
            elif time_step == 'month':
                input_df['day'] = input_df.index.days_in_month
            else:
                pydt = input_df.index[len(input_df) - 1]
                pydt = pd.to_datetime(datetime.datetime(2000, wyem, 1, pydt.hour, pydt.minute))
                pydt = pd.to_datetime(datetime.datetime(pydt.year, pydt.month, pydt.days_in_month, pydt.hour, pydt.minute))
                input_df['day'] = pydt.days_in_month
            input_df['date'] = input_df[['year', 'month', 'day']].apply(
                lambda s : datetime.datetime(*s),axis = 1)
            input_df['date'] = pd.to_datetime(input_df['date'])
            input_df.set_index('date', inplace = True)
        
        # verify period
        
        if start_dt is None:
            pydt = input_df.index[0]
            start_dt = pd.to_datetime(datetime.datetime(pydt.year, pydt.month, pydt.day, pydt.hour, pydt.minute))
        if end_dt is None: 
            pydt = input_df.index[len(input_df) - 1]
            end_dt = pd.to_datetime(datetime.datetime(pydt.year, pydt.month, pydt.day, pydt.hour, pydt.minute))
                
        try:
            input_df = input_df.truncate(before = start_dt, after = end_dt)
        except:
            logging.error('\nERROR: ' + str(sys.exc_info()[0]) + 'occurred truncating input data')
            return return_df
        if len(input_df.index) < 1:
           input_df = make_ts_dataframe(time_step, ts_quantity, start_dt, 
                       end_dt, wyem)
        
        # adjust for missing rows
        
        full_index = make_dt_index(time_step, ts_quantity, start_dt, end_dt, wyem)
        full_index = full_index + pd.Timedelta(full_index[0] - input_df.index[0])
        input_df = input_df.reindex(index = full_index)
    
        # determine values column

        notFound = True
        for column, input_column in enumerate(lc_columns):
            if lc_station in input_column and lc_param in input_column:
                notFound = False
                break
        if notFound:
            logging.error("Unable to locate station " + stationToRead + " and parameter " + parameterToRead + " in file " + file_path + ".")
            return return_df
            
        # merge values
        
        column_name = input_columns[column]
        try:
            return_df = pd.merge(make_ts_dataframe(time_step, ts_quantity, start_dt, end_dt), 
                    input_df[[column_name]], left_index = True, right_index = True)
        except:
            logging.error('\nERROR: ' + str(sys.exc_info()[0]) + 'occurred merging input data with return dataframe.\n')
            # raise
            return return_df
        del input_df, full_index
        return_df = return_df.rename(columns = {column_name:parameterToRead})
        
        # scale values

        return_df[parameterToRead] *= scaleFactor
        return return_df
    except:
        logging.error('\nERROR: ' + str(sys.exc_info()[0]) + ' Error occurred reading column slot data from file\n' + file_path)
        return return_df

def ReadOneTextRDB(file_path, header_lines, names_line, stationToRead, 
        parameterToRead, units, scaleFactor, time_step, ts_quantity, 
        valuesSeparator, start_dt = None, end_dt = None, 
        mia_value = 'NaN', wyem = 12):
    """Reads RDB Text Relational Database (Type C database design) for one station and parameter

     Args:
        file_path: fully specified file path
        header_lines: number of header lines
        names_line: line of header names
        stationToRead: station to read
        parameterToRead: parameter to read
        units: units of parameter
        scaleFactor: scale to apply to parameter
        time_step: RiverWare style string timestep
        ts_quantity: Interger number of time_steps's in interval
        valuesSeparator: separator of values
        start_dt: starting date time
        end_dt: ending date time
        mia_value: missing value
        wyem: Water Year End Month

    Returns:
        populated dataframe
    """
    return_df = None
    lc_station = stationToRead.lower()
    lc_param = parameterToRead.lower()
    try:
        if names_line == 0:
            # default column names and locations
            
            input_df = pd.read_table(file_path, engine = 'python', 
                    header = None, sep = valuesSeparator, na_values = mia_value)
            if input_df.empty:
                logging.error("No data read in file" + file_path)
                return return_df
            sta_column_name = "Station"
            param_column_name = "Parameter"
            date_column_name = "Date"
            values_column_name = "Value"
            input_columns = ['Station', 'Parameter', 'Date', 'Value']
            input_df.columns = input_columns
        else:
            # dynamic column names and location
            
            # Get list of 0 based line numbers to skip - Ignore header but assume header was set as 1's based index
            data_skip = [i for i in range(header_lines) if i + 1 <> names_line]
            input_df = pd.read_table(file_path, engine = 'python',
                    header = names_line - len(data_skip) - 1, skiprows = data_skip, 
                    na_values = mia_value, sep = valuesSeparator)
            if input_df.empty:
                logging.error("No data read in file" + file_path)
                return return_df
            input_columns = list(input_df.columns)
            lc_columns = [x.lower() for x in input_columns]
        
            # determine column types

            try:
                sta_column = lc_columns.index('station')
            except:
                try:
                    sta_column = lc_columns.index('object')
                except:
                    stacolumn = 0
            sta_column_name = input_columns[sta_column]
            try:
                param_column = lc_columns.index('parameter')
            except:
                try:
                    param_column = lc_columns.index('slot')
                except:
                    stacolumn = 1
            param_column_name = input_columns[param_column]
            try:
                date_column = lc_columns.index('date')
            except:
                date_column = 2
            date_column_name = input_columns[date_column]
            try:
                values_column = lc_columns.index('value')
            except:
                values_column = 3
            values_column_name = input_columns[values_column]
        
        # create new column of lower case station and parameter
        
        try:
            input_df["sta_param"] = (input_df[sta_column_name].map(str) + "." 
                                      + input_df[param_column_name].map(str)).str.lower()
        except:
            logging.error('\nERROR: ' + str(sys.exc_info()[0]) + 'occurred creating sta_param column.\n')
            return return_df
    
        # locate requested station and parameter

        sta_params = list(pd.unique(input_df.sta_param.ravel()))
        notFound = True
        for sta_param in sta_params:
            if lc_station in sta_param and lc_param in sta_param:
                notFound = False
                break
        if notFound:
            logging.error("Unable to locate station " + stationToRead + " and parameter " + parameterToRead + " in file " + file_path + ".")
            return return_df
    
        # filter data to requested station and parameter values
        
        try:
            input_df = input_df[input_df.sta_param == sta_param]
        except:
            logging.error('\nERROR: ' + str(sys.exc_info()[0]) + 'occurred filtering to ', sta_param, '.\n')
            return return_df

        # set date column as index

        input_df = input_df.rename(columns = {date_column_name:'date'})
        if time_step == 'year' and len(str(input_df['date'][0])) == 4:
            input_df['date'] = pd.to_datetime(input_df['date'], format = '%Y')
        else:
            input_df['date'] = pd.to_datetime(input_df['date'])
        input_df.set_index('date', inplace = True)
        
        # set starting and ending dates
        
        if start_dt is None:
            pydt = input_df.index[0]
            start_dt = pd.to_datetime(datetime.datetime(pydt.year, pydt.month, pydt.day, pydt.hour, pydt.minute))
        if end_dt is None: 
            pydt = input_df.index[len(input_df) - 1]
            end_dt = pd.to_datetime(datetime.datetime(pydt.year, pydt.month, pydt.day, pydt.hour, pydt.minute))
        try:
            input_df = input_df.truncate(before = start_dt, after = end_dt)
        except:
            logging.error('\nERROR: ' + str(sys.exc_info()[0]) + 'occurred truncating input data')
            return return_df
        
        # adjust for missing rows
        
        full_index = make_dt_index(time_step, ts_quantity, start_dt, end_dt, wyem)
        full_index = full_index + pd.Timedelta(full_index[0] - input_df.index[0])
        input_df = input_df.reindex(index = full_index)

        # merge values
        
        try:
            return_df = pd.merge(make_ts_dataframe(time_step, ts_quantity, start_dt, end_dt), 
                    input_df[[values_column_name]], left_index = True, right_index = True)
        except:
            logging.error('\nERROR: ' + str(sys.exc_info()[0]) + 'occurred merging input data with return dataframe.\n')
            # raise
            return return_df
        del input_df, full_index
        return_df = return_df.rename(columns = {values_column_name:parameterToRead})
        
        # scale values

        return_df[parameterToRead] *= scaleFactor
        return return_df
    except:
        logging.error('\nERROR: ' + str(sys.exc_info()[0]) + ' Error occurred reading RDB data from file\n' + file_path)
        return return_df

def ReadOneExcelColumn(file_path, ws_name, header_lines, names_line, stationToRead, 
        parameterToRead, units, scaleFactor, time_step, ts_quantity, 
        start_dt = None, end_dt = None, mia_value = 'NaN', wyem = 12):
    """Reads one Excel column station and parameter

     Args:
        file_path: fully specified file path
        ws_name: worksheet name
        header_lines: number of header lines
        names_line: line of header names
        stationToRead: station to read
        parameterToRead: parameter to read
        units: units of parameter
        scaleFactor: scale to apply to parameter
        time_step: RiverWare style string timestep
        ts_quantity: Interger number of time_steps's in interval
        valuesSeparator: separator of values
        start_dt: starting date time
        end_dt: ending date time
        mia_value: missing value
        wyem: Water Year End Month

    Returns:
        populated dataframe
    """
    return_df = None
    lc_station = stationToRead.lower()
    lc_param = parameterToRead.lower()
    try:
        # Get list of 0 based line numbers to skip - Ignore header but assume header was set as 1's based index
        data_skip = [i for i in range(header_lines) if i + 1 <> names_line]
        input_df = pd.read_excel(file_path, sheetname = ws_name, index_col = 0,
                    header = names_line  - len(data_skip) - 1, 
                    skiprows = data_skip, na_values = mia_value)
        if input_df.empty:
            logging.error("No data read in file" + file_path)
            return return_df
                
        # deal with excess rows at bottom that show up as NaT dates

        input_df = input_df[pd.notnull(input_df.index)]
            
        # Deal with possible exitence of 23:59 hour:minute in workbook dates from RiverWare dates
            
        input_df.index.names = ['date']
        input_df['year'] = input_df.index.year
        input_df['month'] = input_df.index.month
        input_df['day'] = input_df.index.day
        if time_step == 'hour' or time_step == 'minute':
            input_df['hour'] = input_df.index.hour
            input_df['minute'] = input_df.index.minute
        else:
            input_df['hour'] = 0
            input_df['minute'] = 0
        input_df['Date'] = input_df[['year', 'month', 'day', 'hour', 'minute']].apply(lambda s : datetime.datetime(*s),axis = 1)
        input_df.reset_index('date', inplace = True, drop = True)
        input_df.set_index('Date', inplace = True)
        input_df.drop(['year', 'month', 'day', 'hour', 'minute'], axis = 1, inplace = True)

        input_columns = list(input_df.columns)
        lc_columns = [x.lower() for x in input_columns]
        
        # set starting and ending dates

        if start_dt is None:
            pydt = input_df.index[0]
            if time_step == 'hour' or time_step == 'minute':
                sdt = datetime.datetime(pydt.year, pydt.month, pydt.day, pydt.hour, pydt.minute)
            else:
                sdt = datetime.datetime(pydt.year, pydt.month, pydt.day, 0, 0)
            start_dt = pd.to_datetime(sdt)
        else:
            if time_step == 'day' or time_step == 'month' or time_step == 'year':
                start_dt = pd.to_datetime(datetime.datetime(start_dt.year, start_dt.month, start_dt.day, 0, 0))
        if end_dt is None: 
            pydt = input_df.index[len(input_df) - 1]
            if time_step == 'hour' or time_step == 'minute':
                edt = datetime.datetime(pydt.year, pydt.month, pydt.day, pydt.hour, pydt.minute)
            else:
                edt = datetime.datetime(pydt.year, pydt.month, pydt.day, 23, 59)
            end_dt = pd.to_datetime(edt)
        else:
            if time_step == 'day' or time_step == 'month' or time_step == 'year':
                end_dt = pd.to_datetime(datetime.datetime(end_dt.year, end_dt.month, end_dt.day, 23, 59))
        try:
            input_df = input_df.truncate(before = start_dt, after = end_dt)
        except:
            logging.error('\nERROR: ' + str(sys.exc_info()[0]) + 'occurred truncating input data')
            return return_df
    
        # determine values column

        notFound = True
        for column, input_column in enumerate(lc_columns):
            if lc_station in input_column and lc_param in input_column:
                notFound = False
                break
        if notFound:
            logging.error("Unable to locate station " + stationToRead + " and parameter " + parameterToRead + " in file " + file_path + ".")
            return return_df
        
        # adjust for missing rows
        
        full_index = make_dt_index(time_step, ts_quantity, start_dt, end_dt, wyem)
        full_index = full_index + pd.Timedelta(full_index[0] - input_df.index[0])
        input_df = input_df.reindex(index = full_index)
        column_name = input_columns[column]
            
        # merge values
        
        try:
            return_df = pd.merge(make_ts_dataframe(time_step, ts_quantity, start_dt, end_dt), 
                    input_df[[column_name]], left_index=True, right_index=True)
        except:
            logging.error('\nERROR: ' + str(sys.exc_info()[0]) + 'occurred merging input data with return dataframe.\n')
            # raise
            return return_df
        del input_df, full_index
        return_df = return_df.rename(columns = {column_name:parameterToRead})
        
        # scale values

        return_df[parameterToRead] *= scaleFactor
        return return_df
    except:
        logging.error('\nERROR: ' + str(sys.exc_info()[0]) + ' Error occurred reading data from workbook')
        return return_df

def ColumnSlotToDataframe(file_path, header_lines, names_line, 
        time_step, ts_quantity, valuesSeparator, 
        start_dt = None, end_dt = None, 
        mia_value = 'NaN', wyem = 12):
    """Pull column slot data into a dataframe

     Args:
        file_path: fully specified file path
        header_lines: number of header lines
        names_line: line of header names
        time_step: RiverWare style string timestep
        ts_quantity: Interger number of time_steps's in interval
        valuesSeparator: separator of values
        start_dt: starting date time
        end_dt: ending date time
        mia_value: missing value
        wyem: Water Year End Month

    Returns:
        populated dataframe
    """
    return_df = None
    try:
        # Get list of 0 based line numbers to skip - Ignore header but assume header was set as 1's based index
        data_skip = [i for i in range(header_lines) if i + 1 <> names_line]
        input_df = pd.read_table(file_path, engine = 'python', 
                header = names_line - len(data_skip) - 1, skiprows = data_skip, 
                sep = valuesSeparator, na_values = mia_value)
        if input_df.empty:
            logging.error("No data read in file" + file_path)
            return return_df
        input_columns = list(input_df.columns)
        lc_columns = [x.lower() for x in input_columns]
        
        # determine date column

        try:
            date_column = lc_columns.index('date')
        except:
            date_column = 0
        date_column_name = input_columns[date_column]
        input_df = input_df.rename(columns = {date_column_name:'date'})
        
        # make sure that daily, monthly and annual data use end of period dates and do not include a time stamp
        
        input_df['date'] = pd.to_datetime(input_df['date'])
        input_df.set_index('date', inplace = True)
        if time_step == 'day' or time_step == 'month' or time_step == 'year':
            input_df['year'] = input_df.index.year
            input_df['month'] = input_df.index.month
            if time_step == 'day':
                input_df['day'] = input_df.index.day
            elif time_step == 'month':
                input_df['day'] = input_df.index.days_in_month
            else:
                pydt = input_df.index[len(input_df) - 1]
                pydt = pd.to_datetime(datetime.datetime(2000, wyem, 1, pydt.hour, pydt.minute))
                pydt = pd.to_datetime(datetime.datetime(pydt.year, pydt.month, pydt.days_in_month, pydt.hour, pydt.minute))
                input_df['day'] = pydt.days_in_month
            input_df['date'] = input_df[['year', 'month', 'day']].apply(
                lambda s : datetime.datetime(*s),axis = 1)
            input_df['date'] = pd.to_datetime(input_df['date'])
            input_df.set_index('date', inplace = True)
        
        # verify period
        
        if start_dt is None:
            pydt = input_df.index[0]
            start_dt = pd.to_datetime(datetime.datetime(pydt.year, pydt.month, pydt.day, pydt.hour, pydt.minute))
        if end_dt is None: 
            pydt = input_df.index[len(input_df) - 1]
            end_dt = pd.to_datetime(datetime.datetime(pydt.year, pydt.month, pydt.day, pydt.hour, pydt.minute))

        try:
            input_df = input_df.truncate(before = start_dt, after = end_dt)
        except:
            logging.error('\nERROR: ' + str(sys.exc_info()[0]) + 'occurred truncating input data')
            return return_df
        if len(input_df.index) < 1:
           input_df = make_ts_dataframe(time_step, ts_quantity, start_dt, 
                       end_dt, wyem)
        
        # adjust for missing rows
        
        full_index = make_dt_index(time_step, ts_quantity, start_dt, end_dt, wyem)
        full_index = full_index + pd.Timedelta(full_index[0] - input_df.index[0])
        # could not get user mia_value to work but it appears to work better using default np.nan
        # return_df = input_df.reindex(index = full_index, fill_value = mia_value)
        return_df = input_df.reindex(index = full_index)
        del input_df, full_index
        return return_df
    except:
        logging.error('\nERROR: ' + str(sys.exc_info()[0]) + ' Error occurred reading column slot data from\n' + file_path)
        return return_df

def TextRDBToDataframe(file_path, header_lines, names_line, 
            time_step, ts_quantity, valuesSeparator, 
            start_dt = None, end_dt = None, 
            mia_value = 'NaN', wyem = 12):
    """Reads RDB Text Relational Database (Type C database design) into columnar dataframe

     Args:
        file_path: fully specified file path
        header_lines: number of header lines
        names_line: line of header names
        time_step: RiverWare style string timestep
        ts_quantity: Interger number of time_steps's in interval
        valuesSeparator: separator of values
        start_dt: starting date time
        end_dt: ending date time
        mia_value: missing value
        wyem: Water Year End Month
  
    Returns:
        populated dataframe
    """
    return_df = None
    try:
        if names_line == 0:
            # default column names and locations
            
            input_df = pd.read_table(file_path, engine = 'python', 
                    header = None, sep = valuesSeparator, na_value = mia_value)
            if input_df.empty:
                logging.error("No data read in file" + file_path)
                return return_df
            sta_column_name = "Station"
            param_column_name = "Parameter"
            date_column_name = "Date"
            values_column_name = "Value"
            input_columns = ['Station', 'Parameter', 'Date', 'Value']
            input_df.columns = input_columns
        else:
            # dynamic column names and location
            
            # Get list of 0 based line numbers to skip - Ignore header but assume header was set as 1's based index
            data_skip = [i for i in range(header_lines) if i + 1 <> names_line]
            input_df = pd.read_table(file_path, engine = 'python',
                    header = names_line - len(data_skip) - 1, skiprows = data_skip, 
                    na_values = mia_value, sep = valuesSeparator)
            if input_df.empty:
                logging.error("No data read in file" + file_path)
                return return_df
            input_columns = list(input_df.columns)
            lc_columns = [x.lower() for x in input_columns]
        
            # determine column types

            try:
                sta_column = lc_columns.index('station')
            except:
                try:
                    sta_column = lc_columns.index('object')
                except:
                    stacolumn = 0
            sta_column_name = input_columns[sta_column]
            try:
                param_column = lc_columns.index('parameter')
            except:
                try:
                    param_column = lc_columns.index('slot')
                except:
                    stacolumn = 1
            param_column_name = input_columns[param_column]
            try:
                date_column = lc_columns.index('date')
            except:
                date_column = 2
            date_column_name = input_columns[date_column]
            try:
                values_column = lc_columns.index('value')
            except:
                values_column = 3
            values_column_name = input_columns[values_column]
        
        # create new column of station and parameter
        
        try:
            input_df["sta_param"] = input_df[sta_column_name].map(str) \
                                      + "." + input_df[param_column_name].map(str)
        except:
            logging.error('\nERROR: ' + str(sys.exc_info()[0]) + 'occurred creating sta_param column.\n')
            return return_df
    
        # set starting and ending dates

        input_df = input_df.rename(columns = {date_column_name:'date'})
        if time_step == 'year' and len(str(input_df['date'][0])) == 4:
            input_df['date'] = pd.to_datetime(input_df['date'], format = '%Y')
        else:
            input_df['date'] = pd.to_datetime(input_df['date'])
        pysdt = input_df['date'][0]
        pyedt = input_df['date'][len(input_df) - 1]
        if time_step == 'year':
            adt = pd.to_datetime(datetime.datetime(2000, wyem, 1))
            sdt = datetime.datetime(pysdt.year, wyem, adt.days_in_month, 0, 0)
            edt = datetime.datetime(pyedt.year, wyem, adt.days_in_month, 0, 0)
        elif time_step == 'month':
            sdt = datetime.datetime(pysdt.year, pysdt.month, pysdt.day, 0, 0)
            edt = datetime.datetime(pyedt.year, pyedt.month, pyedt.days_in_month, 0, 0)
        elif time_step == 'day':
            sdt = datetime.datetime(pysdt.year, pysdt.month, pysdt.day, 0, 0)
            edt = datetime.datetime(pyedt.year, pyedt.month, pyedt.day, 0, 0)
        else:
            sdt = datetime.datetime(pysdt.year, pysdt.month, pysdt.day, pysdt.hour, pysdt.minute)
            edt = datetime.datetime(pyedt.year, pyedt.month, pyedt.day, pyedt.hour, pyedt.minute)
        sdt = pd.to_datetime(sdt)
        edt = pd.to_datetime(edt)
        
        # following is going to fail with missing rows but do not know a workaround
        
        return_df = make_ts_dataframe(time_step, ts_quantity, sdt, edt, wyem)
        
        # parse sta_param column into output dataframe

        sta_params = list(pd.unique(input_df.sta_param.ravel()))
        notFound = True
        for sta_param in sta_params:
            try:
                # temp_df = make_ts_dataframe(time_step, ts_quantity, sdt, edt)
                temp_df = input_df[input_df.sta_param == sta_param]
                return_df[sta_param] = temp_df[values_column_name].values
                del temp_df
                # return_df = pd.merge(return_df, temp_df[[sta_param]], left_index = True, right_index = True)
            except:
                logging.error('\nERROR: ' + str(sys.exc_info()[0]) + 'occurred filtering to ', sta_param, '.\n')
                sys.exit()
        del input_df
    
        # set starting and ending dates and truncate final dataframe

        if start_dt is None:
            start_dt = return_df.index[0]
        if end_dt is None: 
            end_dt = return_df.index[len(return_df) - 1]
        try:
            input_df = return_df.truncate(before = start_dt, after = end_dt)
        except:
            logging.error('\nERROR: ' + str(sys.exc_info()[0]) + 'occurred truncating input data')
            return_df = None
            return return_df
        
        # adjust for missing rows

        full_index = make_dt_index(time_step, ts_quantity, start_dt, end_dt, wyem)
        full_index = full_index + pd.Timedelta(full_index[0] - input_df.index[0])
        return_df = input_df.reindex(index = full_index)
        del input_df, full_index
        return return_df
    except:
        logging.error('\nERROR: ' + str(sys.exc_info()[0]) + ' Error occurred reading RDB data from\n' + file_path)
        return return_df

def ExcelWorksheetToDataframe(file_path, ws_name, header_lines, names_line, 
        time_step, ts_quantity, start_dt = None, end_dt = None, 
        mia_value = 'NaN', wyem = 12):
    """Reads one Excel worksheet into a dataframe

     Args:
        file_path: fully specified file path
        ws_name: worksheet name
        header_lines: number of header lines
        names_line: line of header names
        time_step: RiverWare style string timestep
        ts_quantity: Interger number of time_steps's in interval
        valuesSeparator: separator of values
        start_dt: starting date time
        end_dt: ending date time
        mia_value: missing value
        wyem: Water Year End Month

    Returns:
        populated dataframe
    """
    return_df = None
    try:
        # Get list of 0 based line numbers to skip - Ignore header but assume header was set as 1's based index
        data_skip = [i for i in range(header_lines) if i + 1 <> names_line]
        input_df = pd.read_excel(file_path, sheetname = ws_name, index_col = 0,
            header = names_line - len(data_skip) - 1, skiprows = data_skip, na_values = mia_value)
        if input_df.empty:
            logging.error("No data read in file" + file_path)
            return return_df

        # deal with excess rows at bottom that show up as NaT dates

        input_df = input_df[pd.notnull(input_df.index)]
            
        # Deal with possible exitence of 23:59 hour:minute in workbook dates from RiverWare dates
            
        input_df.index.names = ['date']
        input_df['year'] = input_df.index.year
        input_df['month'] = input_df.index.month
        input_df['day'] = input_df.index.day
        if time_step == 'hour' or time_step == 'minute':
            input_df['hour'] = input_df.index.hour
            input_df['minute'] = input_df.index.minute
        else:
            input_df['hour'] = 0
            input_df['minute'] = 0
        input_df['Date'] = input_df[['year', 'month', 'day', 'hour', 'minute']].apply(lambda s : datetime.datetime(*s),axis = 1)
        input_df.reset_index('date', inplace = True, drop = True)
        input_df.set_index('Date', inplace = True)
        input_df.drop(['year', 'month', 'day', 'hour', 'minute'], axis = 1, inplace = True)
        
        # set starting and ending dates

        if start_dt is None:
            pydt = input_df.index[0]
            if time_step == 'hour' or time_step == 'minute':
                sdt = datetime.datetime(pydt.year, pydt.month, pydt.day, pydt.hour, pydt.minute)
            else:
                sdt = datetime.datetime(pydt.year, pydt.month, pydt.day, 0, 0)
            start_dt = pd.to_datetime(sdt)
        else:
            if time_step == 'day' or time_step == 'month' or time_step == 'year':
                start_dt = pd.to_datetime(datetime.datetime(start_dt.year, start_dt.month, start_dt.day, 0, 0))
        if end_dt is None: 
            pydt = input_df.index[len(input_df) - 1]
            if time_step == 'hour' or time_step == 'minute':
                edt = datetime.datetime(pydt.year, pydt.month, pydt.day, pydt.hour, pydt.minute)
            else:
                edt = datetime.datetime(pydt.year, pydt.month, pydt.day, 23, 59)
            end_dt = pd.to_datetime(edt)
        else:
            if time_step == 'day' or time_step == 'month' or time_step == 'year':
                edt = datetime.datetime(end_dt.year, end_dt.month, end_dt.day, 23, 59)
                end_dt = pd.to_datetime(edt)
        try:
            input_df = input_df.truncate(before = start_dt, after = end_dt)
        except:
            logging.error('\nERROR: ' + str(sys.exc_info()[0]) + 'occurred truncating input data')
            return return_df
        
        # adjust for missing rows
        
        full_index = make_dt_index(time_step, ts_quantity, start_dt, end_dt, wyem)
        full_index = full_index + pd.Timedelta(full_index[0] - input_df.index[0])
        # could not get user mia_value to work ????
        # return_df = input_df.reindex(index = full_index, fill_value = mia_value)
        return_df = input_df.reindex(index = full_index)
        del input_df, full_index
        return return_df
    except:
        logging.error('\nERROR: ' + str(sys.exc_info()[0]) + ' Error occurred reading data from workbook')
        return return_df

def ReduceDataframeToParameter(input_df, parameterToUse):
    """Reduces Dataframe to specified parameter
    
     Args:
        input_df: input pandas dataframe
        parameterToUse: parameter to reduce to

    Returns:
        reduced dataframe
    """

    return_df = input_df.copy()
    lc_param = parameterToUse.lower()
    try:
        input_columns = list(input_df.columns)
        for sta_param in input_columns:
            if lc_param not in sta_param.lower():
                return_df.drop(sta_param, axis = 1, inplace = True)
        return return_df
    except:
        logging.error('\nERROR: ' + str(sys.exc_info()[0]) + ' Error occurred reducing to ' + parameterToUse)
        return return_df

def ReadOneDataframeColumn(input_df, stationToRead, parameterToRead, units, 
        scaleFactor, time_step, ts_quantity, start_dt, end_dt):
    """Reads column of a dataframe for specified station and parameter

     Args:
        input_df: input pandas dataframe
        stationToRead: station to read
        parameterToRead: parameter to read
        units: units of parameter
        scaleFactor: scale to apply to parameter
        time_step: RiverWare style string timestep
        ts_quantity: Interger number of time_steps's in interval
        start_dt: starting date time
        end_dt: ending date time

    Returns:
        populated dataframe
    """
    return_df = None
    lc_station = stationToRead.lower()
    lc_param = parameterToRead.lower()
    try:
        input_columns = list(input_df.columns)
        lc_columns = [x.lower() for x in input_columns]
    
        # determine values column

        notFound = True
        for column, input_column in enumerate(lc_columns):
            if lc_station in input_column and lc_param in input_column:
                notFound = False
                break
        if notFound:
            logging.error("Unable to locate station " + stationToRead + " and parameter " + parameterToRead + " in dataframe.")
            return return_df
        column_name = input_columns[column]
            
        # merge values
        
        try:
            return_df = make_ts_dataframe(time_step, ts_quantity, start_dt, end_dt)
            return_df[column_name] = input_df[column_name].values
        except:
            logging.error('\nERROR: ' + str(sys.exc_info()[0]) + 'occurred merging input data with return dataframe.\n')
            # raise
            return return_df
        del input_df
        return_df = return_df.rename(columns = {column_name:parameterToRead})
        
        # scale values

        return_df[parameterToRead] *= scaleFactor
        return return_df
    except:
        logging.error('\nERROR: ' + str(sys.exc_info()[0]) + ' Error occurred reading column data from dataframe')
        return return_df

def csf_output_by_dataframe(file_path, delimiter, new_data_df, 
        float_format, date_format, date_is_posted, 
        mia_value = 'NaN'):
    """Post a DataFrame to a column slot text file

    Args:
        file_path: fully specified file path
        delimiter: delimiter
        new_data_df: new data DataFrame
        float_format: floating point number format
        date_format: date format
        date_is_posted: date is posted flag
        mia_value: missing value

    Returns:
        success: True or False
    """
    logging.debug('  Posting specified data to a text column slot file')
    try:
        if float_format is None:
            if date_is_posted:
                new_data_df.to_csv(path_or_buf = file_path, sep = delimiter, 
                            date_format = date_format, na_rep = mia_value)
            else:
                new_data_df.to_csv(path_or_buf = file_path, sep = delimiter, 
                            index = False, na_rep = mia_value)
        else:    # formatted output causes loss of precision
            if date_is_posted:
                new_data_df.to_csv(path_or_buf = file_path, sep = delimiter, 
                        date_format = date_format, float_format = float_format, 
                        na_rep = mia_value)
            else:
                new_data_df.to_csv(path_or_buf = file_path, sep = delimiter,
                        index = False, float_format = float_format, na_rep = mia_value)
        return True
    except:
        logging.error('\nERROR: ' + str(sys.exc_info()[0]) + 'occurred posting csv output data')
        # raise
        return False

def rdb_output_by_df_nodate_formatting(file_path, 
        delimiter, new_data_df, float_format, 
        date_format, mia_value = 'NaN'):
    """Post a DataFrame to an rdb text file

    Args:
        file_path: fully specified file path
        delimiter: delimiter
        new_data_df: new data DataFrame
        float_format: floating point number format
        date_format: date format
        mia_value: missing value
    """
    logging.debug('  Posting specified data to a text rdb')
    output_dict = {}
    dates_dti = pd.to_datetime(new_data_df.index)
    dates = dates_dti.strftime(date_format)
    try:
        # create dictionary of rdb dataframes
            
        stas = list(new_data_df.columns)
        for sta in stas:
            if '.' in sta:
                split_values = sta.split(".")
                station = split_values[0]
                param = split_values[1]
            else:
                station = sta
                param = 'NaN'
            column_df = pd.DataFrame(columns = ['Station', 'Parameter', 'Date', 'Value'])
            column_df['Date'] = dates
            column_df['Value'] = new_data_df[sta].values
            column_df['Station'] = station
            column_df['Parameter'] = param
            output_dict[sta] = column_df
            del column_df

        # create consolidated rdb dataframe
            
        rdb_df = pd.concat(output_dict)
        if float_format is None:
            rdb_df.to_csv(path_or_buf = file_path, sep = delimiter, 
                        index = False, na_rep = mia_value)
        else:    # formatted output causes loss of precision
            rdb_df.to_csv(path_or_buf = file_path, sep = delimiter,
                    index = False, float_format = float_format, na_rep = mia_value)
        return True
    except:
        logging.error('\nERROR: ' + str(sys.exc_info()[0]) + 'occurred posting rdb output data')
        # raise
        return False
         
def rdb_output_by_dataframe(file_path, delimiter, 
        new_data_df, float_format, date_format, 
        date_is_posted, mia_value = 'NaN'):
    """Post a DataFrame to an rdb text file

    Args:
        file_path: fully specified file path
        delimiter: delimiter
        new_data_df: new data DataFrame
        float_format: floating point number format
        date_format: date format
        date_is_posted: date is posted flag
        mia_value: missing value
    """
    logging.debug('  Posting specified data to a text rdb')
    dates_dti = pd.to_datetime(new_data_df.index)
    dates = dates_dti.strftime(date_format)
    try:
        # create dictionary of rdb dataframes
            
        stas = list(new_data_df.columns)
        for staCount, sta in enumerate(stas):
            if '.' in sta:
                split_values = sta.split(".")
                station = split_values[0]
                param = split_values[1]
            else:
                station = sta
                param = 'NaN'
            column_df = pd.DataFrame(columns = ['Station', 'Parameter', 'Date', 'Value'])
            column_df['Date'] = dates
            column_df['Value'] = new_data_df[sta].values
            column_df['Station'] = station
            column_df['Parameter'] = param
            if float_format is None:
                if staCount == 0:
                    column_df.to_csv(path_or_buf = file_path, sep = delimiter, 
                                index = False, na_rep = mia_value)
                else:
                    column_df.to_csv(path_or_buf = file_path, sep = delimiter, 
                                index = False, na_rep = mia_value,
                                header = False, mode = 'a')
            else:    # formatted output causes loss of precision
                if staCount == 0:
                    column_df.to_csv(path_or_buf = file_path, sep = delimiter,
                        index = False, float_format = float_format, na_rep = mia_value)
                else:
                    column_df.to_csv(path_or_buf = file_path, sep = delimiter,
                        index = False, float_format = float_format, na_rep = mia_value, 
                        header = False, mode = 'a')
            del column_df
        return True
    except:
        logging.error('\nERROR: ' + str(sys.exc_info()[0]) + 'occurred posting rdb output data')
        # raise
        return False

def wb_output_by_df_ew(wb_path, ws_name, new_data_df, 
        float_format, date_format, date_is_posted, time_freq, 
        mia_value = 'NaN'):
    """Post a DataFrame to a workbook given a workbook
       This version is called directrly

    Args:
        wb_path: fully specified workbook path
        ws_name: worksheet name
        new_data_df: new data DataFrame
        float_format: floating point number format
        date_format: date format
        date_is_posted: date is posted flag
        time_freq: 15T, H, 6H, 12H, D, 7D, M, or A
        mia_value: missing value

    Returns:
        success: True or False
    """
    logging.debug('  Posting specified data to a workbook')
    try:
        wb_writer = pd.ExcelWriter(wb_path, datetime_format = date_format,
                    engine = 'xlsxwriter')
        if date_is_posted:
            temp_df = new_data_df.copy().reset_index()
            temp_df.to_excel(wb_writer, ws_name, index = False, 
                    float_format = float_format, na_rep = mia_value)
            del temp_df
        else:
            new_data_df.to_excel(wb_writer, ws_name, index = False, 
                    float_format = float_format, na_rep = mia_value)
        wb = wb_writer.book
        ws = wb_writer.sheets[ws_name]
        format = wb.add_format()
        format.set_text_wrap()
        format.set_bold()
        if 'T' in time_freq: # minute output
            ws.set_column(0, 0, 14)
            ws.set_column(1, len(new_data_df.columns), 14)
        elif 'H' in time_freq: # hourly output
            ws.set_column(0, 0, 14)
            ws.set_column(1, len(new_data_df.columns), 14)
        elif 'D' in time_freq: # daily output
            ws.set_column(0, 0, 12)
            ws.set_column(1, len(new_data_df.columns), 12)
        elif 'M' in time_freq: # post monthly output
            ws.set_column(0, 0, 8)
            ws.set_column(1, len(new_data_df.columns), 11)
        elif 'A' in time_freq: # post annual output
            ws.set_column(0, 0, 8)
            ws.set_column(1, len(new_data_df.columns), 10)
        else:
            ws.set_column(0, 0, 12)
            ws.set_column(1, len(new_data_df.columns), 12)
        for column, column_name in enumerate(list(new_data_df.columns)):
            if date_is_posted:
                ws.write_string(0, column + 1, column_name, format)
            else:
                ws.write_string(0, column, column_name, format)
        ws.set_selection(1, 1, 1, 1)
        ws.freeze_panes(1, 1)
        wb_writer.save()
        del wb_writer, wb, ws
        return True
    except:
        logging.error('\nERROR: ' + str(sys.exc_info()[0]) + 'occurred posting workbook output data')
        return False
         
def wb_output_by_df_xlsxwriter(wb_writer, wb, ws_name, new_data_df, 
        float_format, date_is_posted, time_freq, 
        mia_value = 'NaN'):
    """Post a DataFrame to a workbook given an ExcelWriter object
       This version if called by wb_output_via_df_dict_xlsxwriter

    Args:
        wb_writer: pandas ExcelWriter object
        wb: xlsxwriter Workbook oject
        ws_name: worksheet name
        new_data_df: new data DataFrame
        float_format: floating point number format
        date_is_posted: date is posted flag
        time_freq: 15T, H, 6H, 12H, D, 7D, M, or A
        mia_value: missing value

    Returns:
        success: True or False
    """
    try:
        if date_is_posted:
            temp_df = new_data_df.copy().reset_index()
            temp_df.to_excel(wb_writer, ws_name, index = False, 
                    float_format = float_format, na_rep = mia_value)
            del temp_df
        else:
            new_data_df.to_excel(wb_writer, ws_name, index = False, 
                    float_format = float_format, na_rep = mia_value)
        ws = wb_writer.sheets[ws_name]
        header_format = wb.add_format()
        header_format.set_text_wrap()
        header_format.set_bold()
        if 'T' in time_freq: # minute output
            ws.set_column(0, 0, 14)
            ws.set_column(1, len(new_data_df.columns), 14)
        elif 'H' in time_freq: # hourly output
            ws.set_column(0, 0, 14)
            ws.set_column(1, len(new_data_df.columns), 14)
        elif 'D' in time_freq: # daily output
            ws.set_column(0, 0, 12)
            ws.set_column(1, len(new_data_df.columns), 12)
        elif 'M' in time_freq: # post monthly output
            ws.set_column(0, 0, 8)
            ws.set_column(1, len(new_data_df.columns), 11)
        elif 'A' in time_freq: # post annual output
            ws.set_column(0, 0, 8)
            ws.set_column(1, len(new_data_df.columns), 10)
        else:
            ws.set_column(0, 0, 12)
            ws.set_column(1, len(new_data_df.columns), 12)
        for column, column_name in enumerate(list(new_data_df.columns)):
            if date_is_posted:
                ws.write_string(0, column + 1, column_name, header_format)
            else:
                ws.write_string(0, column, column_name, header_format)
        ws.set_selection(1, 1, 1, 1)
        ws.freeze_panes(1, 1)
        del ws
        return True
    except:
        logging.error('\nERROR: ' + str(sys.exc_info()[0]) + 'occurred posting ExcelWriter output data')
        return False

def wb_output_via_df_dict_xlsxwriter(wb_path, ws_names, new_data_dict, 
        float_format, date_format, date_is_posted, 
        time_step, ts_quantity, mia_value = 'NaN', 
        replace_flag = True, wyem = 12, date_formats = []):
    """Post a dictionary of DataFrame's to a workbook
       
       xlswriter can only write to Excel.  Therefore,
       to write to existing workbook and retain existing content
       such as worksheets other than one being written to,
       code has to rebuild existing workbook with merged
       content.  This version include ability to optinally
       merge existing data of worksheet being posted
       to with new data.  In addition, formulas are lost.
       
       List date_formats was created to support not losing
       date formatting of existing time series worksheets.
       If a calling application does not provide date_formats
       list, defaults are used based on worksheet names.
       
       Worksheet names are also used to guess what worksheets
       to treat as time series worksheets and which to
       consider non time series worksheets. No attempt 
       is made to preserve formatting of
       existing non time series worksheets.

       replace_flag tells code to replace all rows if True.
       If replace_flag is False, rows are merged temporally.
       Code always attempts to merge columns when posting
       to an existing worksheet.

    Args:
        wb_path: fully specified workbook path
        ws_names: list of worksheet names
        new_data_dict: dictionary of data frames to post
        float_format: floating point number format
        date_format: date format
        date_is_posted: date is posted flag
        time_step: RiverWare style string timestep
        ts_quantity: Interger number of time_steps's in interval
        mia_value: missing value
        replace_flag: True or False
        wyem: Water Year End Month
        date_formats: All available date formats

    Returns:
        success: True or False
    """
    if len(date_formats) < 1:
        date_formats.append('m/d/yyyy h:mm')    # minute
        date_formats.append('m/d/yyyy h:mm')    # hourly
        date_formats.append('m/d/yyyy')    # daily
        date_formats.append('m/d/yyyy')    # weekly
        date_formats.append('m/yyyy')    # monthly
        date_formats.append('m/yyyy')    # annul
    try:
        time_freq = get_ts_freq(time_step, ts_quantity, wyem)
        wb_writer = pd.ExcelWriter(wb_path, datetime_format = date_format,
                    engine = 'xlsxwriter')
        wb = wb_writer.book
        if os.path.isfile(wb_path):
            # workbook exists - copy and write to an ExcelWriter object
        
            # existing_sheets_dict = pd.read_excel(wb_path, sheetname = None, na_values = ['NaN'])
            existing_sheets_dict = pd.read_excel(wb_path, sheetname = None, na_values = mia_value)
            header_format = wb.add_format()
            header_format.set_text_wrap()
            header_format.set_bold()
    
            # add existing content of workbook to wb_writer

            for sn, existing_df in existing_sheets_dict.items():
                # check if a time series worksheet
        
                existing_columns = list(existing_df.columns)
                existing_lc_cols = [x.lower() for x in existing_columns]
                if existing_lc_cols[0] == 'date':
                    if sn in ws_names:
                        existing_df.set_index(existing_columns[0], inplace = True)
                        existing_df.index.names = ['date']
                        existing_df['year'] = existing_df.index.year
                        existing_df['month'] = existing_df.index.month
                        existing_df['day'] = existing_df.index.day
                        if 'T' in time_freq or 'H' in time_freq:
                            existing_df['hour'] = existing_df.index.hour
                            existing_df['minute'] = existing_df.index.minute
                        else:
                            existing_df['hour'] = 0
                            existing_df['minute'] = 0
                        existing_df['Date'] = existing_df[['year', 'month', 'day', 'hour', 'minute']].apply(lambda s : datetime.datetime(*s),axis = 1)
                        existing_df.reset_index('date', inplace = True, drop = True)
                        existing_df.set_index('Date', inplace = True)
                        existing_df.drop(['year', 'month', 'day', 'hour', 'minute'], axis = 1, inplace = True)
                        sheet_date_format = date_format

                        # new data exist in a common worksheet name - merge data

                        key_index = ws_names.index(sn)
                        key_name = new_data_dict.keys()[key_index]
                        new_data_df = new_data_dict[key_name]
                        if replace_flag:
                            # replace existing period with new period
                            
                            existing_df = existing_df.reindex(index = new_data_df.index, fill_value = 'NaN')
                        else:
                            # blend existing and new periods
                            # assuming that RiverWare type date times already exist in existing and new

                            pydt = min(existing_df.index[0], new_data_df.index[0])
                            start_dt = datetime.datetime(pydt.year, pydt.month, pydt.day, pydt.hour, pydt.minute)
                            pydt = max(existing_df.index[len(existing_df) - 1], new_data_df.index[len(new_data_df) - 1])
                            end_dt = datetime.datetime(pydt.year, pydt.month, pydt.day, pydt.hour, pydt.minute)
                            try:
                                existing_df = existing_df.truncate(before = start_dt, after = end_dt)
                            except:
                                logging.error('\nERROR: ' + str(sys.exc_info()[0]) + 'occurred truncating input data')
                                return False
                            existing_df = existing_df.reindex(index = pd.date_range(start_dt, end_dt, 
                                        freq = time_freq, name = "Date"), fill_value = 'NaN')
                            new_data_df = new_data_df.reindex(index = pd.date_range(start_dt, end_dt, 
                                        freq = time_freq, name = "Date"), fill_value = 'NaN')
                        del new_data_dict[key_name]
                        ws_names.remove(sn)

                        """
                        # unable to get merge or join to work
                        try:
                            existing_df = pd.merge(new_data_df, existing_df, left_index = True, right_index = True)
                        except:
                            logging.error('\nERROR: ' + str(sys.exc_info()[0]) + 'occurred merging new data with existing data.\n')
                            return False
                        """
                        # brut force merging approach

                        new_columns = list(new_data_df.columns)
                        
                        # pick up new values for existing columns
                        
                        for col_name in existing_columns:
                            if col_name in new_columns:
                                if replace_flag:
                                    existing_df[col_name] = new_data_df[col_name].values
                                else:
                                    existing_values = existing_df[col_name].values
                                    new_data_values = new_data_df[col_name].values
                                    for v_count in range(0, len(new_data_values)):
                                        if not new_data_values[v_count] is None and not new_data_values[v_count] == 'NaN':
                                            existing_values[v_count] = new_data_values[v_count]
                                    existing_df[col_name] = existing_values
                                    del existing_values, new_data_values, v_count
                        existing_columns = list(existing_df.columns)
                        
                        # add new columns
                        
                        for col_name in new_columns:
                            if not col_name in existing_columns:
                                existing_df[col_name] = new_data_df[col_name].values
                    else:    # estimate date time format to use
                        existing_df.set_index(existing_columns[0], inplace = True)
                        try:
                            lcsn = sn.lower()
                            if 'instant' in lcsn or 'min' in lcsn:
                                sheet_date_format = date_formats[0]
                            elif 'hour' in lcsn:
                                sheet_date_format = date_formats[1]
                            elif 'daily' in lcsn or 'day' in lcsn:
                                sheet_date_format = date_formats[2]
                            elif 'week' in lcsn:
                                sheet_date_format = date_formats[3]
                            elif 'month' in lcsn:
                                sheet_date_format = date_formats[4]
                            elif 'year' in lcsn or 'annual' in lcsn:
                                sheet_date_format = date_formats[5]
                            elif 'day' in lcsn:
                                sheet_date_format = date_formats[2]
                            else:
                                sheet_date_format = date_format
                            sheet_date_format.replace('Y', 'yyyy').replace('%H','h').replace('M', 'mm')
                        except:
                            sheet_date_format = date_format
                    wb_writer.datetime_format = sheet_date_format
                    if not wb_output_by_df_xlsxwriter(wb_writer, wb, 
                            sn, existing_df, float_format, date_is_posted, 
                            time_freq, mia_value):
                        del wb_writer, wb
                        return False
                else:    # non time series worksheet
                    existing_df.to_excel(wb_writer, sn, index = False, na_rep = mia_value)
                    ws = wb_writer.sheets[sn]
                    for column, column_name in enumerate(list(existing_df.columns)):
                        if 'Unname' in column_name:
                            ws.write_string(0, column, "", header_format)
                        else:
                            ws.write_string(0, column, column_name, header_format)
            del existing_sheets_dict
        if len(new_data_dict.keys()) > 0:
            field_count = -1
            for field_name, new_data_df in new_data_dict.items():
                field_count += 1
                if not wb_output_by_df_xlsxwriter(wb_writer, wb, 
                        ws_names[field_count], new_data_df, float_format, 
                        date_is_posted, time_freq, mia_value):
                    del wb_writer, wb
                    return False
        wb_writer.save()
        del wb_writer, wb
        return True
    except:
        logging.error('\nERROR: ' + str(sys.exc_info()[0]) + 'occurred posting workbook data from dictionary')
        return False

def wb_output_via_df_dict_openpyxl(wb_path, ws_names, new_data_dict, 
        float_format, date_format, time_step, ts_quantity, 
        mia_value = 'NaN', replace_flag = True, wyem = 12):
    """Post a dictionary of DataFrame's to a workbook
       
       openpyxl can read and write to Excel and
       has more capability than xlsxwriter.  Futhermore,
       it is more consistent with open office concepts.
       This code is able to perserve formulas of worksheets
       that are not being posted to.  In addition,
       because this code is able to write to an existing
       workbook without having to recreate workbook,
       it runs much faster than xlswriter version.
       
       Formulas on non posted worksheets are posted
       without calculated values (an openpyxl behavoir).
       If values are not visible after opening a workbook,
       press 'F9' to calculate formulas and save workvook.
       
       replace_flag tells code to replace all rows if True.
       If replace_flag is False, rows are merged temporally.
       Code always attempts to merge columns when posting
       to an existing worksheet.

    Args:
        wb_path: fully specified workbook path
        ws_names: list of worksheet names
        new_data_dict: dictionary of data frames to post
        float_format: floating point number format
        date_format: date format
        time_step: RiverWare style string timestep
        ts_quantity: Interger number of time_steps's in interval
        mia_value: missing value
        replace_flag: True or False
        wyem: Water Year End Month

    Returns:
        success: True or False
    """
    try:
        logging.debug('Reading existing content of workbook\n' + wb_path)
        time_freq = get_ts_freq(time_step, ts_quantity, wyem)
        if not os.path.isfile(wb_path):
            # workbook does not exist - create an empty one

            wb = op.Workbook()
            wb.save(filename = wb_path)
            new_wb = True
        else:
            new_wb = False
            
        # open one copy of workbook for reading and one for writing
        
        existing_wb = op.load_workbook(wb_path, data_only = True)
        existing_sheets = existing_wb.sheetnames
        for ws_name in existing_sheets:
            if not ws_name in ws_names:
                existing_wb.remove_sheet(existing_wb[ws_name])
        existing_sheets = existing_wb.sheetnames
        wb = op.load_workbook(wb_path)
        
        # loop thru list of worksheets to post
        
        for ws_name in ws_names:
            # set up worksheet
            
            if ws_name in existing_sheets:
                wb.remove_sheet(wb[ws_name])
            ws = wb.create_sheet(title = ws_name)
            key_index = ws_names.index(ws_name)
            key_name = new_data_dict.keys()[key_index]
            new_data_df = new_data_dict[key_name]
            if ws_name in existing_sheets:
                # merge new data with existing data

                try:
                    dv = existing_wb[ws_name].values
                    dc = next(dv)
                    existing_df = pd.DataFrame(data = dv, columns = dc) # works
                except:
                    logging.error('\nERROR: ' + str(sys.exc_info()[0]) + 'occurred pulling existing data from workbook')
                    sys.exit()
                existing_columns = list(existing_df.columns)
                
                # deal with excess rows at bottom that show up as NaT dates

                existing_df = existing_df[pd.notnull(existing_df[existing_columns[0]])]
                existing_df.set_index(existing_columns[0], inplace = True)
                existing_df.index.names = ['date']
                existing_df['year'] = existing_df.index.year
                existing_df['month'] = existing_df.index.month
                existing_df['day'] = existing_df.index.day
                if 'T' in time_freq or 'H' in time_freq:
                    existing_df['hour'] = existing_df.index.hour
                    existing_df['minute'] = existing_df.index.minute
                else:
                    existing_df['hour'] = 0
                    existing_df['minute'] = 0
                existing_df['Date'] = existing_df[['year', 'month', 'day', 'hour', 'minute']].apply(lambda s : datetime.datetime(*s),axis = 1)
                existing_df.reset_index('date', inplace = True, drop = True)
                existing_df.set_index('Date', inplace = True)
                existing_df.drop(['year', 'month', 'day', 'hour', 'minute'], axis = 1, inplace = True)
                existing_columns = list(existing_df.columns)
                if replace_flag:
                    # replace existing period with new period

                    existing_df = existing_df.reindex(index = new_data_df.index, fill_value = 'NaN')
                else:
                    # blend existing and new periods
                    # assuming that RiverWare type date times already exist in existing and new
                
                        pydt = min(existing_df.index[0], new_data_df.index[0])
                        start_dt = datetime.datetime(pydt.year, pydt.month, pydt.day, pydt.hour, pydt.minute)
                        pydt = max(existing_df.index[len(existing_df) - 1], new_data_df.index[len(new_data_df) - 1])
                        end_dt = datetime.datetime(pydt.year, pydt.month, pydt.day, pydt.hour, pydt.minute)
                        try:
                            existing_df = existing_df.truncate(before = start_dt, after = end_dt)
                        except:
                            logging.error('\nERROR: ' + str(sys.exc_info()[0]) + 'occurred truncating input data')
                            return False
                        existing_df = existing_df.reindex(index = pd.date_range(start_dt, end_dt, 
                                    freq = time_freq, name = "Date"), fill_value = 'NaN')
                        new_data_df = new_data_df.reindex(index = pd.date_range(start_dt, end_dt, 
                                    freq = time_freq, name = "Date"), fill_value = 'NaN')
                # brut force merging approach

                new_columns = list(new_data_df.columns)
                        
                # pick up new values for existing columns
                        
                for col_name in existing_columns:
                    if col_name in new_columns:
                        existing_values = existing_df[col_name].values
                        new_data_values = new_data_df[col_name].values
                        for v_count in range(0, len(new_data_values)):
                            if not new_data_values[v_count] is None and not new_data_values[v_count] == 'NaN':
                                existing_values[v_count] = new_data_values[v_count]
                        existing_df[col_name] = existing_values
                        del existing_values, new_data_values, v_count
                existing_columns = list(existing_df.columns)
                        
                # add new columns
                        
                for col_name in new_columns:
                    if not col_name in existing_columns:
                        existing_df[col_name] = new_data_df[col_name].values
                new_data_df = existing_df.copy()
                del existing_df

            # post data

            for row in dataframe_to_rows(new_data_df, index = True, header = True):
                ws.append(row)
            for row in ws.iter_cols(min_col = 2, max_row = 1):
                for cur_cell in row:
                    ws.column_dimensions[op.utils.cell.get_column_letter(cur_cell.col_idx)].width = 10.00
            for row in ws.iter_cols(max_row = 1):
                for cur_cell in row:
                    cur_cell.alignment =  op.styles.alignment.Alignment(wrap_text = True)
                    cur_cell.border =  op.styles.borders.Border(outline = op.styles.borders.Side(border_style = None))
                    cur_cell.font =  op.styles.Font(bold = True)
            if date_format is not None:
                for row in ws.iter_rows(max_col = 1):
                    for cur_cell in row:
                        cur_cell.number_format = date_format
                        cur_cell.font =  op.styles.Font(bold = True)
            if float_format is not None:
                for row in ws.iter_rows(min_row = 2, min_col = 2):
                    for cur_cell in row:
                        cur_cell.number_format = float_format
            cur_cell = ws.cell(row = 1, column = 1)
            cur_cell.value = "Date"
            if 'T' in time_freq: # minute output
                column_width = 16.0
            elif 'H' in time_freq: # hourly output
                column_width = 16.0
            elif 'D' in time_freq: # daily output
                column_width = 12.0
            elif 'M' in time_freq: # post monthly output
                column_width = 11.0
            elif 'A' in time_freq: # post annual output
                column_width = 10.0
            else:
                column_width = 12.0
            ws.column_dimensions['A'].width = column_width
            cur_cell.border =  op.styles.borders.Border(outline = op.styles.borders.Side(border_style = 'thin'))
            ws.freeze_panes = 'B2'
        if new_wb:
            wb.remove_sheet(wb['Sheet'])
        wb.save(wb_path)
        del wb, existing_wb
        return True
    except:
        logging.error('\nERROR: ' + str(sys.exc_info()[0]) + 'occurred posting workbook data from dictionary')
        sys.exit()
